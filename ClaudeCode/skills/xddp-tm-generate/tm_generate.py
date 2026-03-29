"""
変更設計書からTMを自動生成するスクリプト。
xddp-tm-generate Skillの実行例として使用する。
"""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── 変更設計書のパース ────────────────────────
def parse_design_doc(filepath):
    """変更設計書からエントリ一覧を抽出する"""
    with open(filepath, encoding="utf-8") as f:
        content = f.read()

    entries = []
    # 「## 変更エントリ：」セクションを全件抽出
    pattern = r"## 変更エントリ：(.+?)\n(.*?)(?=\n## 変更エントリ：|\n## 変更履歴|$)"
    for m in re.finditer(pattern, content, re.DOTALL):
        header_line = m.group(1).strip()  # 例: CRS-2024-031-02-01 × sensor_common.h
        body = m.group(2)

        # 仕様番号を抽出（複数の場合はカンマ区切り）
        spec_ids = []
        spec_match = re.search(r"\| 仕様番号\s*\|\s*(.+?)\s*\|", body)
        if spec_match:
            raw = spec_match.group(1).strip()
            spec_ids = [s.strip() for s in re.split(r"[,、]", raw)]

        # ソースファイルを抽出
        file_match = re.search(r"\| ソースファイル\s*\|\s*(.+?)\s*\|", body)
        src_file = file_match.group(1).strip() if file_match else ""

        # 変更対象を抽出
        target_match = re.search(r"\| 変更対象\s*\|\s*(.+?)\s*\|", body)
        target = target_match.group(1).strip() if target_match else ""

        # 修正方針を抽出（変更ファイルサマリ用）
        policy_match = re.search(r"### 修正方針\n(.+?)(?=\n###)", body, re.DOTALL)
        policy = policy_match.group(1).strip()[:60] + "..." if policy_match else ""

        # 修正エントリかどうかを判定（テストNGフィードバック由来）
        is_fix = "NG対応" in header_line

        for spec_id in spec_ids:
            entries.append({
                "spec_id": spec_id,
                "src_file": src_file,
                "target": target,
                "policy": policy,
                "is_fix": is_fix,
                "mark": "●",
            })

    return entries


# ── スペックアウト資料のパース ─────────────────
def parse_specout(filepath):
    """スペックアウト資料から○の交点候補を抽出する"""
    refs = []
    try:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()

        # 呼び出し関係テーブルから影響有無「有」「確認中」を抽出
        table_match = re.search(
            r"### 呼び出し関係テーブル\n(.+?)(?=\n###|\n---)", content, re.DOTALL
        )
        if table_match:
            for row in table_match.group(1).strip().split("\n"):
                cols = [c.strip() for c in row.split("|") if c.strip()]
                if len(cols) >= 5 and cols[4] in ["有", "確認中"]:
                    # 呼び出し先の関数からファイルは推測できないため記録のみ
                    refs.append({
                        "caller": cols[0],
                        "callee": cols[1],
                        "impact": cols[4],
                    })
    except FileNotFoundError:
        pass
    return refs


# ── 変更要求仕様書のパース ────────────────────
def parse_crs(filepath):
    """変更要求仕様書から仕様番号→カテゴリ・仕様概要のマップを生成する"""
    spec_map = {}
    try:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()

        current_category = "機能"
        category_pattern = r"### カテゴリ：(.+)"
        spec_pattern = r"\| (CRS-[\d-]+) \| [^\|]+ \| [^\|]+ \| [^\|]+ \| [^\|]+ \| (.+?) \| (.+?) \|"

        for line in content.split("\n"):
            cat_m = re.match(category_pattern, line)
            if cat_m:
                current_category = cat_m.group(1).strip()

            spec_m = re.match(spec_pattern, line)
            if spec_m:
                spec_id = spec_m.group(1).strip()
                before  = spec_m.group(2).strip()
                after   = spec_m.group(3).strip()
                spec_map[spec_id] = {
                    "category": current_category,
                    "summary": f"{before} → {after}",
                }
    except FileNotFoundError:
        pass
    return spec_map


# ── TM構築 ───────────────────────────────────
def build_tm(entries, spec_map):
    """エントリリストからTMデータ構造を構築する"""
    # ソースファイル一覧（出現順を維持）
    files = []
    for e in entries:
        if e["src_file"] and e["src_file"] not in files:
            files.append(e["src_file"])

    # 仕様番号一覧（出現順を維持）
    spec_ids = []
    for e in entries:
        if e["spec_id"] and e["spec_id"] not in spec_ids:
            spec_ids.append(e["spec_id"])

    # 交点マトリクス {spec_id: {file: mark}}
    matrix = {s: {f: "" for f in files} for s in spec_ids}
    for e in entries:
        if e["spec_id"] in matrix and e["src_file"] in matrix[e["spec_id"]]:
            matrix[e["spec_id"]][e["src_file"]] = e["mark"]

    # 変更ファイルサマリ
    file_summary = {}
    for f in files:
        change_count = sum(1 for s in spec_ids if matrix[s][f] == "●")
        ref_count    = sum(1 for s in spec_ids if matrix[s][f] == "○")
        policy_list  = [e["policy"] for e in entries if e["src_file"] == f and e["mark"] == "●"]
        file_summary[f] = {
            "change": change_count,
            "ref": ref_count,
            "policy": policy_list[0] if policy_list else "-",
        }

    return files, spec_ids, matrix, file_summary


# ── Excel出力 ────────────────────────────────
def write_excel(files, spec_ids, matrix, file_summary, spec_map, cr_id, title, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "TM"

    FONT_NAME = "Arial"
    C_HEADER = "1F4E79"
    C_META   = "F5F5F5"
    C_HEAD   = "2E75B6"
    C_FILE   = "D9E8F5"
    C_MARK   = "1F4E79"
    C_REF    = "70AD47"
    C_CAT = {
        "機能": "D6E4F0", "テスト": "FFF2CC",
        "日程": "FCE4D6", "成果物管理": "EAD1DC",
    }

    def fill(c):
        return PatternFill("solid", fgColor=c)
    def border():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)
    def al(h="left", v="center", wrap=True, rot=0):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap, text_rotation=rot)

    # 列幅
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 34
    for i in range(len(files)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 18

    last_col = get_column_letter(3 + len(files))
    row = 1

    # タイトル
    ws.merge_cells(f"A{row}:{last_col}{row}")
    ws[f"A{row}"] = "トレーサビリティマトリクス（TM）【自動生成】"
    ws[f"A{row}"].font = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
    ws[f"A{row}"].fill = fill(C_HEADER)
    ws[f"A{row}"].alignment = al("center")
    ws.row_dimensions[row].height = 26
    row += 1

    # メタ
    for label, value in [
        ("対応CR番号", cr_id),
        ("タイトル", title),
        ("生成方法", "xddp-tm-generate Skillにより自動生成"),
    ]:
        ws.merge_cells(f"A{row}:B{row}")
        ws.merge_cells(f"C{row}:{last_col}{row}")
        ws[f"A{row}"] = label
        ws[f"C{row}"] = value
        ws[f"A{row}"].font = Font(name=FONT_NAME, bold=True, size=9)
        ws[f"C{row}"].font = Font(name=FONT_NAME, size=9)
        ws[f"A{row}"].fill = fill(C_META)
        ws[f"C{row}"].fill = fill(C_META)
        ws[f"A{row}"].border = border()
        ws[f"C{row}"].border = border()
        ws[f"A{row}"].alignment = al()
        ws[f"C{row}"].alignment = al()
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1

    # 凡例
    ws.merge_cells(f"A{row}:{last_col}{row}")
    ws[f"A{row}"] = "凡例：● 変更あり（ソースファイルを変更する）　○ 参照・影響確認あり（変更なし）　空白 関係なし"
    ws[f"A{row}"].font = Font(name=FONT_NAME, size=9, color="444444")
    ws[f"A{row}"].fill = fill("FFFDE7")
    ws[f"A{row}"].border = border()
    ws[f"A{row}"].alignment = al()
    ws.row_dimensions[row].height = 16
    row += 1

    row += 1

    # ヘッダ行1段目
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"] = "仕様"
    ws.merge_cells(f"D{row}:{last_col}{row}")
    ws[f"D{row}"] = "ソースファイル"
    for col, label in [("A","仕様"), ("D","ソースファイル")]:
        ws[f"{col}{row}"].font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        ws[f"{col}{row}"].fill = fill(C_HEADER)
        ws[f"{col}{row}"].border = border()
        ws[f"{col}{row}"].alignment = al("center")
    ws.row_dimensions[row].height = 16
    row += 1

    # ヘッダ行2段目
    header_row = row
    for col, h in [("A","カテゴリ"),("B","仕様番号"),("C","仕様概要")]:
        ws[f"{col}{row}"] = h
        ws[f"{col}{row}"].font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        ws[f"{col}{row}"].fill = fill(C_HEAD)
        ws[f"{col}{row}"].border = border()
        ws[f"{col}{row}"].alignment = al("center")
    for i, fname in enumerate(files):
        c = get_column_letter(4 + i)
        ws[f"{c}{row}"] = fname
        ws[f"{c}{row}"].font = Font(name=FONT_NAME, bold=True, size=9, color="FFFFFF")
        ws[f"{c}{row}"].fill = fill(C_FILE)
        ws[f"{c}{row}"].border = border()
        ws[f"{c}{row}"].alignment = al("center", rot=45)
    ws.row_dimensions[row].height = 60
    row += 1

    # データ行
    for idx, spec_id in enumerate(spec_ids):
        info = spec_map.get(spec_id, {"category": "-", "summary": "-"})
        cat   = info["category"]
        bg    = "FFFFFF" if idx % 2 == 0 else "F2F8FF"
        cat_c = C_CAT.get(cat, bg)

        ws[f"A{row}"] = cat
        ws[f"A{row}"].font = Font(name=FONT_NAME, bold=True, size=9)
        ws[f"A{row}"].fill = fill(cat_c)
        ws[f"A{row}"].border = border()
        ws[f"A{row}"].alignment = al("center")

        ws[f"B{row}"] = spec_id
        ws[f"B{row}"].font = Font(name=FONT_NAME, size=9)
        ws[f"B{row}"].fill = fill(bg)
        ws[f"B{row}"].border = border()
        ws[f"B{row}"].alignment = al()

        ws[f"C{row}"] = info["summary"]
        ws[f"C{row}"].font = Font(name=FONT_NAME, size=9)
        ws[f"C{row}"].fill = fill(bg)
        ws[f"C{row}"].border = border()
        ws[f"C{row}"].alignment = al()

        for i, fname in enumerate(files):
            c = get_column_letter(4 + i)
            mark = matrix[spec_id][fname]
            ws[f"{c}{row}"] = mark
            ws[f"{c}{row}"].border = border()
            ws[f"{c}{row}"].alignment = al("center")
            if mark == "●":
                ws[f"{c}{row}"].font = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
                ws[f"{c}{row}"].fill = fill(C_MARK)
            elif mark == "○":
                ws[f"{c}{row}"].font = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
                ws[f"{c}{row}"].fill = fill(C_REF)
            else:
                ws[f"{c}{row}"].fill = fill(bg)

        ws.row_dimensions[row].height = 22
        row += 1

    # サマリ
    row += 1
    ws.merge_cells(f"A{row}:{last_col}{row}")
    ws[f"A{row}"] = "変更ファイルサマリ"
    ws[f"A{row}"].font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = fill(C_HEADER)
    ws[f"A{row}"].alignment = al("center")
    ws.row_dimensions[row].height = 20
    row += 1

    ws.merge_cells(f"B{row}:C{row}")
    ws.merge_cells(f"F{row}:{last_col}{row}")
    for col, h in [("A","ファイル名"),("B","変更仕様数"),("D","参照確認数"),("F","変更概要")]:
        ws[f"{col}{row}"] = h
        ws[f"{col}{row}"].font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        ws[f"{col}{row}"].fill = fill(C_HEAD)
        ws[f"{col}{row}"].border = border()
        ws[f"{col}{row}"].alignment = al("center")
    ws.row_dimensions[row].height = 16
    row += 1

    for i, fname in enumerate(files):
        s = file_summary[fname]
        bg = "FFFFFF" if i % 2 == 0 else "F2F8FF"
        ws[f"A{row}"] = fname
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"] = s["change"]
        ws[f"D{row}"] = s["ref"]
        ws.merge_cells(f"E{row}:{last_col}{row}")
        ws[f"E{row}"] = s["policy"]
        for col in ["A","B","D","E"]:
            ws[f"{col}{row}"].font = Font(name=FONT_NAME, size=9)
            ws[f"{col}{row}"].fill = fill(bg)
            ws[f"{col}{row}"].border = border()
            ws[f"{col}{row}"].alignment = al()
        ws[f"B{row}"].alignment = al("center")
        ws[f"D{row}"].alignment = al("center")
        ws.row_dimensions[row].height = 18
        row += 1

    ws.freeze_panes = f"A{header_row+1}"
    wb.save(out_path)


# ── メイン ───────────────────────────────────
if __name__ == "__main__":
    DESIGN_DOC  = "/mnt/user-data/outputs/変更設計書-CR-2024-031.md"
    SPECOUT_DOC = "/mnt/user-data/outputs/スペックアウト-CR-2024-031.md"
    CRS_DOC     = "/mnt/user-data/outputs/変更要求仕様書-CRS-2024-031.md"
    OUT_XLSX    = "/home/claude/TM-CR-2024-031-auto.xlsx"

    print("変更設計書を解析中...")
    entries  = parse_design_doc(DESIGN_DOC)
    spec_map = parse_crs(CRS_DOC)

    print(f"  抽出エントリ数: {len(entries)}")
    for e in entries:
        print(f"  {e['spec_id']} × {e['src_file']} → {e['mark']}")

    files, spec_ids, matrix, file_summary = build_tm(entries, spec_map)

    print(f"\nTM構築完了:")
    print(f"  仕様番号: {len(spec_ids)}件")
    print(f"  ソースファイル: {len(files)}本")
    total_mark  = sum(1 for s in spec_ids for f in files if matrix[s][f] == "●")
    total_ref   = sum(1 for s in spec_ids for f in files if matrix[s][f] == "○")
    print(f"  ●: {total_mark}箇所")
    print(f"  ○: {total_ref}箇所")

    write_excel(files, spec_ids, matrix, file_summary, spec_map,
                "CR-2024-031", "温度センサー異常検知のしきい値変更", OUT_XLSX)
    print(f"\nExcel生成完了: {OUT_XLSX}")
