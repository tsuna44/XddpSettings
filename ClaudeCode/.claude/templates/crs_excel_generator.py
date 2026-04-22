"""
CRS Excel Generator — XDDP 変更要求仕様書 Excel 生成スクリプト
Usage: python crs_excel_generator.py <output_path> [--version 1.0]

Expected row structure (6 columns: A–F):

  UR 3行セット (D9E1F2, bold):
    Row1: A=【ユーザ要求】  B=UR-x  C=title  D=''  E=''  F=''
    Row2: A=''  B=理由  C=reason  D=''  E=''  F=''
    Row3: A=''  B=説明  C=explanation  D=''  E=''  F=''

  SR 3行セット (E7E6E6, bold):
    Row1: A=【システム要求】  B=''  C=SR-x-y  D=title  E=''  F=''
    Row2: A=''  B=''  C=理由  D=reason  E=''  F=''
    Row3: A=''  B=''  C=説明  D=explanation  E=''  F=''

  SP title  (F5F5F5, normal):
    A=''  B=''  C=SP-title  D=''  E=''  F=''

  SP Before (FFF2CC, normal):
    A=【仕様】  B=''  C=SP-x-y.z  D='■ Before'  E=before  F=ステータス

  SP After  (E2EFDA, normal):
    A=''  B=''  C=''  D='■ After'  E=after  F=''

  SP 備考 (FFF2CC, normal):  ※備考が空の場合は省略
    A=''  B=''  C=''  D='■ 備考'  E=備考内容  F=''

  SP 懸念・検討事項 (FCE4D6, normal):  ※懸念が空の場合は省略
    A=''  B=''  C=''  D='■ 懸念・検討事項'  E=懸念内容  F=''

  未決事項ヘッダ (BDD7EE, bold):
    A='■ 未決事項'  B='#'  C='項目'  D='内容'  E='対応期限'  F=''

  未決事項データ (DDEBF7, normal):
    A=''  B=Q#  C=項目名  D=内容  E=対応期限  F=''

  気づきヘッダ (C6EFCE, bold):
    A='■ 気づき・提案メモ'  B='#'  C='種別'  D='内容'  E='対応方針'  F=''

  気づきデータ (EBF1DE, normal):
    A=''  B=#  C=種別  D=内容  E=対応方針  F=''

Column widths: A=14, B=13, C=32, D=48, E=37.5, F=18(ステータス)
Row heights: header=20, UR/SR title row=40, reason/explanation rows=auto, SP-title=auto, SP-Before/After=45
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Palette ──────────────────────────────────────────────────────────────────
C_HEADER  = "4472C4"
C_UR      = "D9E1F2"
C_SR      = "E7E6E6"
C_SP      = "F5F5F5"
C_BEFORE  = "FFF2CC"
C_AFTER   = "E2EFDA"
C_BIKO    = "FFF2CC"  # SP 備考（Before と同色）
C_KENEN   = "FCE4D6"  # SP 懸念・検討事項（薄オレンジ）
C_PEND_H  = "BDD7EE"  # 未決事項ヘッダ（薄青）
C_PEND    = "DDEBF7"  # 未決事項データ（極薄青）
C_NOTES_H = "C6EFCE"  # 気づき・提案メモヘッダ（薄緑）
C_NOTES   = "EBF1DE"  # 気づき・提案メモデータ（極薄緑）

def _fill(hex6): return PatternFill("solid", fgColor=hex6)
def _al(wrap=True): return Alignment(horizontal="left", vertical="top", wrap_text=wrap, indent=0)
def _bdr():
    t = Side(style="thin", color="BBBBBB")
    return Border(left=t, right=t, top=t, bottom=t)

def _cell(ws, row, col, value, color, bold, row_h=None, wrap=True):
    c = ws.cell(row=row, column=col)
    c.value     = value
    c.fill      = _fill(color)
    c.font      = Font(bold=bold, color="FFFFFF" if color == C_HEADER else "000000")
    c.alignment = _al(wrap=wrap)
    c.border    = _bdr()
    if row_h is not None:
        ws.row_dimensions[row].height = row_h
    return c

def _row(ws, row, values_colors_bolds, row_h=None, wrap=True):
    """Write one full row. values_colors_bolds = list of (value, color, bold)."""
    for col, (val, clr, bld) in enumerate(values_colors_bolds, 1):
        _cell(ws, row, col, val, clr, bld, row_h if col == 1 else None, wrap=wrap)


# ── Public API ────────────────────────────────────────────────────────────────

def add_header_row(ws, row=1):
    labels = ["レベル", "ID", "内容", "理由・説明 / Before・After", "", "ステータス"]
    _row(ws, row, [(l, C_HEADER, True) for l in labels], row_h=20)


def add_ur_row(ws, row, ur_id, title, reason, explanation="", status=""):
    """UR 3行セット（縦配置）: タイトル行 + 理由行 + 説明行。次の行番号を返す。"""
    _row(ws, row,
         [("【ユーザ要求】", C_UR, True),
          (ur_id,           C_UR, True),
          (title,           C_UR, True),
          ("",              C_UR, True),
          ("",              C_UR, True),
          (status or "",    C_UR, True)],
         row_h=40)
    _row(ws, row + 1,
         [("",           C_UR, True),
          ("理由",       C_UR, True),
          (reason or "", C_UR, True),
          ("",           C_UR, True),
          ("",           C_UR, True),
          ("",           C_UR, True)])
    _row(ws, row + 2,
         [("",                  C_UR, True),
          ("説明",              C_UR, True),
          (explanation or "",   C_UR, True),
          ("",                  C_UR, True),
          ("",                  C_UR, True),
          ("",                  C_UR, True)])
    return row + 3


def add_sr_row(ws, row, sr_id, title, reason, explanation="", status=""):
    """SR 3行セット（縦配置）: タイトル行 + 理由行 + 説明行。次の行番号を返す。"""
    _row(ws, row,
         [("【システム要求】", C_SR, True),
          ("",          C_SR, True),
          (sr_id,       C_SR, True),
          (title,       C_SR, True),
          ("",          C_SR, True),
          (status or "", C_SR, True)],
         row_h=40)
    _row(ws, row + 1,
         [("",           C_SR, True),
          ("",           C_SR, True),
          ("理由",       C_SR, True),
          (reason or "", C_SR, True),
          ("",           C_SR, True),
          ("",           C_SR, True)])
    _row(ws, row + 2,
         [("",                  C_SR, True),
          ("",                  C_SR, True),
          ("説明",              C_SR, True),
          (explanation or "",   C_SR, True),
          ("",                  C_SR, True),
          ("",                  C_SR, True)])
    return row + 3


def add_sp_rows(ws, start_row, sp_id, title, before, after, biko="", kenen="", status=""):
    """SP 3〜5行セット: タイトル行 + Before行 + After行 + (備考行) + (懸念行)。次の行番号を返す。"""
    r = start_row

    # SP title row (F5F5F5) — wrap_text=False で改行させない
    _row(ws, r,
         [("",      C_SP, False),
          ("",      C_SP, False),
          (title,   C_SP, False),
          ("",      C_SP, False),
          ("",      C_SP, False),
          ("",      C_SP, False)],
         wrap=False)
    r += 1

    # Before row (FFF2CC)
    _row(ws, r,
         [("【仕様】",       C_BEFORE, False),
          ("",               C_BEFORE, False),
          (sp_id,            C_BEFORE, False),
          ("■ Before",       C_BEFORE, False),
          (before or "",     C_BEFORE, False),
          (status or "",     C_BEFORE, False)],
         row_h=45)
    r += 1

    # After row (E2EFDA)
    _row(ws, r,
         [("",               C_AFTER, False),
          ("",               C_AFTER, False),
          ("",               C_AFTER, False),
          ("■ After",        C_AFTER, False),
          (after or "",      C_AFTER, False),
          ("",               C_AFTER, False)],
         row_h=45)
    r += 1

    # 備考 row — 内容がある場合のみ出力 (FFF2CC)
    if biko:
        _row(ws, r,
             [("",           C_BIKO, False),
              ("",           C_BIKO, False),
              ("",           C_BIKO, False),
              ("■ 備考",     C_BIKO, False),
              (biko,         C_BIKO, False),
              ("",           C_BIKO, False)])
        r += 1

    # 懸念・検討事項 row — 内容がある場合のみ出力 (FCE4D6)
    if kenen:
        _row(ws, r,
             [("",                   C_KENEN, False),
              ("",                   C_KENEN, False),
              ("",                   C_KENEN, False),
              ("■ 懸念・検討事項",   C_KENEN, False),
              (kenen,                C_KENEN, False),
              ("",                   C_KENEN, False)])
        r += 1

    return r


def add_pending_section(ws, row, items):
    """未決事項セクション。items = list of (number, title, content, deadline)。次の行番号を返す。"""
    r = row
    # Section header
    _row(ws, r,
         [("■ 未決事項", C_PEND_H, True),
          ("#",           C_PEND_H, True),
          ("項目",        C_PEND_H, True),
          ("内容",        C_PEND_H, True),
          ("対応期限",    C_PEND_H, True),
          ("",            C_PEND_H, True)],
         row_h=20)
    r += 1
    for num, title, content, deadline in items:
        _row(ws, r,
             [("",          C_PEND, False),
              (str(num),    C_PEND, False),
              (title or "", C_PEND, False),
              (content or "",C_PEND, False),
              (deadline or "",C_PEND, False),
              ("",          C_PEND, False)])
        r += 1
    return r


def add_notes_section(ws, row, items):
    """気づき・提案メモセクション。items = list of (number, kind, content, policy)。次の行番号を返す。"""
    r = row
    # Section header
    _row(ws, r,
         [("■ 気づき・提案メモ", C_NOTES_H, True),
          ("#",                  C_NOTES_H, True),
          ("種別",               C_NOTES_H, True),
          ("内容",               C_NOTES_H, True),
          ("対応方針",           C_NOTES_H, True),
          ("",                   C_NOTES_H, True)],
         row_h=20)
    r += 1
    for num, kind, content, policy in items:
        _row(ws, r,
             [("",           C_NOTES, False),
              (str(num),     C_NOTES, False),
              (kind or "",   C_NOTES, False),
              (content or "",C_NOTES, False),
              (policy or "", C_NOTES, False),
              ("",           C_NOTES, False)])
        r += 1
    return r


def set_column_widths(ws):
    ws.column_dimensions['A'].width = 14.0
    ws.column_dimensions['B'].width = 13.0
    ws.column_dimensions['C'].width = 32.0
    ws.column_dimensions['D'].width = 48.0
    ws.column_dimensions['E'].width = 37.5
    ws.column_dimensions['F'].width = 18.0  # ステータス列


def add_history_sheet(wb, history_rows):
    """
    history_rows: list of (version, date, author, description)
    """
    ws = wb.create_sheet("変更履歴")
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 65

    for col, h in enumerate(["版数", "日付", "変更者", "変更内容"], 1):
        c = ws.cell(row=1, column=col)
        c.value = h
        c.fill  = _fill(C_HEADER)
        c.font  = Font(bold=True, color="FFFFFF")
        c.alignment = _al()
        c.border = _bdr()
    ws.row_dimensions[1].height = 20

    for ri, (ver, date, author, desc) in enumerate(history_rows, 2):
        for col, val in enumerate([ver, date, author, desc], 1):
            c = ws.cell(row=ri, column=col)
            c.value = val
            c.fill  = _fill(C_SP)
            c.font  = Font(bold=False, color="000000")
            c.alignment = _al()
            c.border = _bdr()
        ws.row_dimensions[ri].height = 50


# ── Example: CRS-CR-2026-001 v2.2 ────────────────────────────────────────────

def build_crs_cr2026_001(out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "変更要求仕様書"
    set_column_widths(ws)

    r = 1
    add_header_row(ws, r); r += 1

    # ── 機能要求 ──────────────────────────────────────────────────────────────
    # UR-001
    r = add_ur_row(ws, r, "UR-001", "タスク追加時に優先度を指定したい",
        "タスク生成時点で重要度を明示することで、タスク一覧での並び替えやフィルタリングの基準となり、"
        "ユーザが重要なタスクを即座に把握できるようにするため",
        explanation="優先度の値は high, medium, low の3段階とする（UR-002と関連）",
        status="確定")

    r = add_sr_row(ws, r, "SR-001-001", "Task オブジェクトに priority 属性を付与する",
        "タスクに優先度情報を保持するために、Task オブジェクトおよびデータ永続化層が priority 属性を"
        "記録・管理できる状態にする必要があるため",
        explanation="SR-001-001 はタスク追加時の優先度受け取り・保存の全般的な要件。具体的な CLI インターフェースは SR-001-002 で定義",
        status="確定")

    r = add_sp_rows(ws, r, "SP-001-001.001", "Task モデルへの priority フィールド追加",
        "Task オブジェクトに priority 属性が存在しない",
        "Task オブジェクトに `priority: str = 'medium'` フィールドを追加する。許容値は \"high\", \"medium\", \"low\" に限定する",
        biko="`add` コマンドで `--priority` が省略された場合は argparse の `default='medium'` により自動的に `medium` が渡される（SP-002-001.001 と連携）。tasks.json への書き込み時に priority フィールドを JSON キーとして保存する",
        status="確定")

    r = add_sp_rows(ws, r, "SP-001-001.002", "tasks.json への priority フィールド保存",
        "tasks.json に priority キーが存在しない",
        "新規作成タスクの JSON 記録に `\"priority\": \"high\" | \"medium\" | \"low\"` キーを含めて保存する",
        biko="`Task.to_dict()` の戻り値に priority キーを追加することで実現する。storage.py 側の変更は不要（Section 4 参照）。",
        status="確定")

    r = add_sr_row(ws, r, "SR-001-002", "`add` コマンドに `--priority` オプションを追加する",
        "ユーザが新規タスク追加時に priority を明示的に指定するための CLI インターフェースが必要なため",
        explanation="`add` コマンドは既存で機能しており、`--priority` オプションはオプショナル（UR-002 でデフォルト値を定義）",
        status="確定")

    r = add_sp_rows(ws, r, "SP-001-002.001", "`add` コマンドのオプション仕様",
        "`add <タイトル>` コマンドは priority 属性なしでタスクを生成する",
        "`add <タイトル> [--priority {high|medium|low}]` をサポートする。`--priority` オプション省略時はデフォルト値 `medium` を使用する（UR-002）",
        biko="argparse にて `--priority` オプションを定義。`choices=['high', 'medium', 'low']` でバリデーション。オプション形式は `--priority high` の空白区切り（bash 標準）",
        status="確定")

    # UR-002
    r = add_ur_row(ws, r, "UR-002", "優先度未指定時にデフォルト値 medium を自動適用してほしい",
        "省略時の既定動作を定めることで、既存スクリプト・ワークフローの互換性を保ちながら新機能を導入できるようにするため",
        explanation="デフォルト値を medium に統一することで、ユーザが何も指定しない場合も一貫性のある優先度付けが実現される",
        status="確定")

    r = add_sr_row(ws, r, "SR-002-001", "`--priority` 省略時に priority のデフォルト値 medium を適用する",
        "`add` コマンドで `--priority` が省略された場合、または既存タスクのロード時に priority フィールドが欠落している場合、"
        "medium をデフォルト値として一貫して適用する必要があるため",
        status="確定")

    r = add_sp_rows(ws, r, "SP-002-001.001", "`add` コマンド実行時のデフォルト値設定",
        "`add` コマンドは `--priority` 省略時に priority 属性を持たないタスクを生成する",
        "`add` コマンドで `--priority` が省略された場合、priority を自動的に `medium` として設定する",
        biko="argparse の `default='medium'` 設定で実装。UI/ログに明示する必要なし（内部デフォルト処理）",
        status="確定")

    r = add_sp_rows(ws, r, "SP-002-001.002", "既存タスクデータの後方互換読み込み",
        "priority フィールドが存在しないタスク JSON レコードは priority 属性が存在しない状態でロードされる",
        "tasks.json 読み込み時に priority フィールドが欠落しているタスクは、priority を `medium` として扱う",
        biko="storage.py の JSON 読み込みロジックで `dict.get('priority', 'medium')` パターンを使用。スキーママイグレーション不要（遅延デフォルト処理）",
        kenen="既存 priority なしデータが存在する状態で `done` コマンド実行後もフィールドが保持されることを回帰テストで確認",
        status="確定")

    # UR-003
    r = add_ur_row(ws, r, "UR-003", "優先度でタスク一覧をフィルタしたい",
        "指定した優先度のタスクのみを表示することで、重要タスク・低優先タスク等の分類ビューが実現でき、"
        "ユーザの関心に応じた情報フォーカスが可能になるため",
        explanation="SR-003-001 の `list` コマンドオプション追加に対応",
        status="確定")

    r = add_sr_row(ws, r, "SR-003-001", "`list` コマンドに `--priority` フィルタオプションを追加する",
        "ユーザが優先度でタスク検索・絞り込みを行うための CLI インターフェースが必要なため",
        status="確定")

    r = add_sp_rows(ws, r, "SP-003-001.001", "`list` コマンドの priority フィルタオプション仕様",
        "`list` コマンドは全タスクを priority の区別なく表示する",
        "`list [--priority {high|medium|low}]` をサポートする。`--priority` 指定時は、該当する priority を持つタスクのみ出力する。未指定時は全タスクを出力する（既存動作変更なし）",
        biko="argparse にて `--priority` オプションを定義。`choices=['high', 'medium', 'low']` でバリデーション。`--priority` 未指定時の出力フォーマットは既存 `list` 動作に従う（後方互換）",
        kenen="Q1（priority をリスト出力に表示カラムとして含めるか、またはフィルタのみか）：フォールバック方針として priority 表示フォーマットは現在の `list` コマンド出力に上乗せせず、フィルタのみを先行実装。priority 表示の要否は次版以降に委ねる",
        status="確定")

    # UR-004
    r = add_ur_row(ws, r, "UR-004", "優先度と status を同時にフィルタしたい",
        "優先度と完了状態を組み合わせた絞り込みにより、「高優先度で未完了」「低優先度で完了済み」等、"
        "複数条件での検索が可能になり、ユーザの作業効率が向上するため",
        explanation="UR-003 の拡張。status フィルタの既存実装はスペックアウトで確認済み（positional 引数 `list [todo|done]`）",
        status="確定")

    r = add_sr_row(ws, r, "SR-004-001", "priority と status を AND 条件でフィルタする",
        "`list` コマンドで priority と status が同時に指定された場合、両方の条件に合致するタスクのみ出力する必要があるため",
        explanation="status フィルタは positional 引数（`list [todo|done]`）として main.py / cli.py に実装済みであることをスペックアウトで確認",
        status="確定")

    r = add_sp_rows(ws, r, "SP-004-001.001", "`list` コマンドの複数オプション AND 動作",
        "`list` コマンドは単一条件フィルタのみ対応する（status フィルタが positional 引数として既存実装）",
        "`list` コマンド実行時に `--priority` と positional status 引数（`todo`|`done`）の両方が指定されたとき、両方の条件に合致するタスクのみを表示する（AND フィルタ）",
        biko="AND フィルタは逐次的に適用（priority 条件で絞った結果に status 条件を適用等）。両オプション未指定時は全タスク表示（UR-NF-001 後方互換）。片方のオプションのみ指定の場合は該当条件でのみフィルタ。argparse 移行後も `nargs='?'` で positional status 引数（`list [todo|done]`）を維持する方針を DSN-CR-2026-001 にて決定",
        status="確定")

    # UR-005
    r = add_ur_row(ws, r, "UR-005", "既存タスクデータの後方互換性を維持したい",
        "priority フィールドなし既存タスクを読み込む際、デフォルト値で自動補完することで、"
        "既存ユーザのデータ資産を無変換で継続利用できるようにするため",
        explanation="旧バージョンの tasks.json から新バージョンへの移行時、priority フィールドが未設定のレコードを自動的に medium で補完する仕組みを提供",
        status="確定")

    r = add_sr_row(ws, r, "SR-005-001", "priority フィールドのない既存データを後方互換で読み込む",
        "旧バージョンの tasks.json（priority フィールドなし）を新バージョンで読み込む際、"
        "priority を `medium` として一貫して扱う仕組みが必要なため",
        explanation="データ永続化層（storage.py）での読み込み処理を指す",
        status="確定")

    r = add_sp_rows(ws, r, "SP-005-001.001", "既存タスクデータの後方互換読み込み",
        "priority フィールドが存在しないタスク JSON レコードは priority 属性が存在しない状態でロードされる",
        "tasks.json 読み込み時に priority フィールドが欠落しているタスクは、priority を `medium` として扱う",
        biko="SP-002-001.002 と同一実装（`Task.from_dict()` 内の `dict.get('priority', 'medium')` パターン）。UR-002（デフォルト値）と UR-005（後方互換性）は同一仕様で実現される。",
        status="確定")

    # ── 非機能要求 ────────────────────────────────────────────────────────────
    # UR-NF-001
    r = add_ur_row(ws, r, "UR-NF-001", "既存コマンドの互換性を維持したい",
        "priority 機能を追加する際、既存ユーザのワークフロー・スクリプト連携が破損しないようにするため。"
        "既存スクリプトが依存する `add` / `list` / `done` / `delete` コマンドの引数・出力・終了コードが変化してはいけない",
        status="確定")

    r = add_sr_row(ws, r, "SR-NF-001-001", "既存コマンドの動作・インターフェースを変更しない（制約）",
        "`--priority` オプション追加後も既存コマンドが同じインターフェース・出力を保つことで、"
        "既存スクリプト・自動化連携の継続性が保証されるため",
        status="確定")

    r = add_sp_rows(ws, r, "SP-NF-001-001.001", "既存コマンドの引数・出力・終了コード維持",
        "`add` / `list` / `done` / `delete` コマンドは既存仕様で動作する",
        "priority オプション追加後も、既存の引数形式・出力フォーマット・終了コードを維持する。"
        "`add <タイトル>` 形式での呼び出しが正常に動作し（priority = medium）、`list` 呼び出しで全タスクを表示する。"
        "`done` / `delete` コマンドは priority 機能追加後も既存の引数・出力・終了コードを維持する",
        biko="既存スクリプトが `add タスク1` と呼び出したとき priority = medium でタスク作成される。"
             "既存スクリプトが `list` と呼び出したとき全タスク表示される。"
             "既存スクリプトが `done <タスクID>` と呼び出したとき該当タスクが完了状態に変わり priority フィールドは維持される",
        status="確定")

    # UR-NF-002
    r = add_ur_row(ws, r, "UR-NF-002", "不正な優先度値を入力できないようにしたい",
        "priority 値が high / medium / low 以外の不正な値で汚染されるのを防ぎ、"
        "データ整合性・システム動作の予測可能性を保証するため",
        explanation="バリデーション失敗時の詳細挙動は SP-NF-002-001.001 で定義",
        status="確定")

    r = add_sr_row(ws, r, "SR-NF-002-001", "`--priority` オプションの入力値を検証する",
        "`add` / `list` コマンドで `--priority` が指定される全箇所で、"
        "許容値（high / medium / low）以外を拒否する仕組みが必要なため",
        status="確定")

    r = add_sp_rows(ws, r, "SP-NF-002-001.001", "優先度値のバリデーション仕様",
        "priority 値の検証なし（データベースに任意の値が入る可能性がある）",
        "`add` / `list` コマンドで `--priority` オプションに high / medium / low 以外の値が指定されたとき、"
        "エラーメッセージを標準エラー出力に表示してコマンドを終了コード 1 で終了する",
        biko="argparse の `choices=['high', 'medium', 'low']` で許容値を制限。"
             "argparse のデフォルトエラーメッセージ例：`error: argument --priority: invalid choice: 'invalid' (choose from high, medium, low)`。"
             "エラー出力先：stderr。終了コード：1。argparse デフォルトエラーメッセージを採用（DSN-CR-2026-001 にて決定）。カスタムメッセージは次回 CR 候補",
        status="確定")

    # ── 未決事項 (Section 5) ──────────────────────────────────────────────────
    r = add_pending_section(ws, r, [
        (1, "priority 表示形式",
         "[フォールバック方針適用・次回CR持ち越し・クローズ - 2026-04-22] `list` コマンド出力への priority カラム追加は本CR スコープ外。フィルタのみ先行実装し、表示フォーマットは次版以降に委ねる（SP-003-001.001 備考参照）。",
         "クローズ済み"),
        (2, "status フィルタ既存実装",
         "[解決済み - 2026-04-16] status フィルタは positional 引数（`list [todo|done]`）として main.py / cli.py に実装済みであることをスペックアウトで確認。SP-004-001.001 修正済み。",
         "解決済み"),
        (3, "エラーメッセージ文言",
         "[解決済み - 2026-04-16] argparse デフォルトエラーメッセージを採用（DSN-CR-2026-001 にて決定）。",
         "解決済み"),
        (4, "priority ソート",
         "要求書 6.3「あれば良い」に priority ソート機能の記載あり。[スコープ外・クローズ - 2026-04-22] 本 CR では priority フィルタのみを対象とし、ソート機能は次回 CR に持ち越す。",
         "クローズ済み"),
    ])

    # ── 気づき・提案メモ (Section 6) ─────────────────────────────────────────
    r = add_notes_section(ws, r, [
        (1, "質問",
         "Q1（priority 表示形式）が未回答のまま CRS 進行すると、設計段階で後戻りが発生する可能性。CRS 確定前に回答取得を強く推奨",
         "今回対応（確認待ち）"),
        (2, "改善案",
         "argparse の `choices` 引数でバリデーション（UR-NF-002）を自動化可能。低コストで実装できるため、優先的に採用推奨",
         "今回対応（設計時に検討）"),
        (3, "改善案",
         "Task クラスに priority を追加する際、`Enum` クラスを使用すると型安全性が向上。標準ライブラリのみで実現可能",
         "今回対応（設計時に検討）"),
        (4, "テスト懸念",
         "`done` コマンド実行後、priority フィールドが保持されることを回帰テストで確認が必須（UR-NF-001 検証の一部）",
         "テスト設計段階で対応"),
        (5, "スコープ外",
         "既存タスクの priority 変更（`update` / `edit` コマンド）は本 CR スコープ外。必要性が認識されれば次 CR で起票",
         "保留"),
        (6, "スコープ確認",
         "priority ソート（要求書 6.3）をこのCR内で取り込むか、次回以降かを依頼者と確認",
         "次回CR（Q4 クローズ済み、本CR スコープ外確定）"),
    ])

    # ── 変更履歴 (Section 7) ──────────────────────────────────────────────────
    add_history_sheet(wb, [
        ("1.0", "2026-04-10", "AI（xddp-spec-writer-agent）",
         "初版作成。ANA-CR-2026-001 セクション2の分類結果に基づき、9個のUR（7個の機能要求 + 2個の非機能要求）に対応する SR/SP を定義"),
        ("1.1", "2026-04-10", "AI（xddp-spec-writer-agent）",
         "レビュー指摘対応（CRT-001: UR-005 を独立 UR として追加、CRT-002: UR 要件数を6個から7個に修正、CRT-003: Q1・Q2 フォールバック方針を明記）"),
        ("1.2", "2026-04-16", "AI（xddp-spec-writer-agent）",
         "スペックアウト結果反映。Section 1 models.py → task.py 修正、Section 4 影響範囲更新（storage.py 変更不要確認）、SP-004-001.001 status フィルタ positional 引数に修正、Q2 解決済みにクローズ"),
        ("1.3", "2026-04-16", "AI（xddp-spec-writer-agent）",
         "DSN-CR-2026-001 決定内容反映。SP-NF-002-001.001 ステータスを「確定」に更新（Q3 解決：argparse デフォルトエラーメッセージ採用）、SP-NF-001-001.001 After の `--done` 誤記を positional 引数表記に修正"),
        ("1.4", "2026-04-16", "AI（xddp-spec-writer-agent）",
         "CHD-CR-2026-001 v1.1 Section 8 の設計フィードバック反映。SP-001-001.001 備考を priority デフォルト値付き定義に修正（DSN 案A 採用と整合）。TM の設計列に CHD-CR-2026-001 の対応セクションを記入"),
        ("2.0", "2026-04-22", "AI（xddp-spec-writer-agent）",
         "ID体系を正規化。UR-XXX/SR-XXX-YYY/SP-XXX-YYY.ZZZ（3桁ゼロパディング）に統一。見出しレベルをテンプレート準拠（h5/h6/h7）に修正。カテゴリ・要求グループ・仕様グループ見出しを追加。内容は v1.4 を踏襲"),
        ("2.1", "2026-04-22", "AI（xddp-spec-writer-agent）",
         "AIレビュー Round 1 指摘対応（#1: SP-005-001.001 に SP-002-001.002 参照を明記し重複を整理、#2: Before 表現3箇所を能動形に修正、#3: SP-001-001.002 備考に Task.to_dict() による実現を明記、#4: Q4 priority ソートをスコープ外クローズ宣言、#7: TM に UR-005 行を追加、#8: SP-NF-001-001.001 After の否定形を正の表現に修正）"),
        ("2.2", "2026-04-22", "AI（xddp-spec-writer-agent）",
         "AIレビュー Round 2 指摘対応（N1: TM の UR-005 重複行を削除、N2: Q1 をクローズ宣言、N3: Section 6 項番6 の対応方針を Q4 クローズ決定と整合）"),
    ])

    wb.save(out_path)
    print(f"Saved: {out_path}  (total rows: {r - 1})")


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "CRS-output.xlsx"
    build_crs_cr2026_001(out)
