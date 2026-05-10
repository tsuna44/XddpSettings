"""
crs_md2excel.py — XDDP 変更要求仕様書 Markdown → Excel 変換スクリプト

Usage: python crs_md2excel.py <CRS_MD_PATH> <OUTPUT_XLSX_PATH>

Expected CRS Markdown structure:
  ## 2. USDM 要求仕様
    ##### UR-xxx タイトル          (or UR-NF-xxx for backward compat)
      ###### SR-xxx-yyy タイトル   (or SR-NF-xxx-yyy)
        ####### SP-xxx-yyy.zzz タイトル  (or SP-NF-xxx-yyy.zzz)
  ## 5. 未決事項          (Markdown table)
  ## 6. 気づき・提案メモ  (Markdown table)
  ## 付記A. スコープ外事項              (Markdown table, optional)
  ## 付記B. 前提条件・実装参考情報      (Markdown table, optional)
  ## 7. 変更履歴          (Markdown table → 変更履歴 sheet)

Excel row structure (6 columns: A–F):
  UR 3行セット (D9E1F2, bold):
    Row1: A=【ユーザ要求】  B=UR-x  C=title  D=''  E=''  F=status
    Row2: A=''  B=理由  C=reason  ...
    Row3: A=''  B=説明  C=explanation  ...

  SR 3行セット (E7E6E6, bold):
    Row1: A=【システム要求】  B=''  C=SR-x-y  D=title  ...  F=status
    Row2: A=''  B=''  C=理由  D=reason  ...
    Row3: A=''  B=''  C=説明  D=explanation  ...

  SP title  (F5F5F5, normal):  A–F = ''  C=title
  SP Before (FFF2CC): A=【仕様】  C=SP-x-y.z  D=■ Before  E=before  F=status
  SP After  (E2EFDA): D=■ After  E=after
  SP 備考   (FFF2CC): D=■ 備考   E=biko  (省略可)
  SP 懸念   (FCE4D6): D=■ 懸念・検討事項  E=kenen  (省略可)

Column widths: A=14, B=13, C=32, D=48, E=37.5, F=18
"""

import os
import re
import sys
from dataclasses import dataclass, field
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Palette ──────────────────────────────────────────────────────────────────
C_HEADER  = "4472C4"
C_UR      = "D9E1F2"
C_SR      = "E7E6E6"
C_SP      = "F5F5F5"
C_BEFORE  = "FFF2CC"
C_AFTER   = "E2EFDA"
C_BIKO    = "FFF2CC"
C_KENEN   = "FCE4D6"
C_PEND_H  = "BDD7EE"
C_PEND    = "DDEBF7"
C_NOTES_H    = "C6EFCE"
C_NOTES      = "EBF1DE"
C_SCOPE_H    = "F4CCCC"
C_SCOPE      = "FCF4F4"
C_IMPL_REF_H = "FFE5B4"
C_IMPL_REF   = "FFF8EC"

def _fill(hex6): return PatternFill("solid", fgColor=hex6)
def _al(wrap=True): return Alignment(horizontal="left", vertical="top", wrap_text=wrap, indent=0)
def _bdr():
    t = Side(style="thin", color="BBBBBB")
    return Border(left=t, right=t, top=t, bottom=t)

def _cell(ws, row, col, value, color, bold, row_h=None, wrap=True):
    c = ws.cell(row=row, column=col)
    c.value     = value
    c.fill      = _fill(color)
    c.font      = Font(bold=bold, color="FFFFFF" if color == C_HEADER else "000000")
    c.alignment = _al(wrap=wrap)
    c.border    = _bdr()
    if row_h is not None:
        ws.row_dimensions[row].height = row_h
    return c

def _row(ws, row, values_colors_bolds, row_h=None, wrap=True):
    """Write one full row. values_colors_bolds = list of (value, color, bold)."""
    for col, (val, clr, bld) in enumerate(values_colors_bolds, 1):
        _cell(ws, row, col, val, clr, bld, row_h if col == 1 else None, wrap=wrap)


# ── Public API ────────────────────────────────────────────────────────────────

def add_header_row(ws, row=1):
    labels = ["レベル", "ID", "内容", "理由・説明 / Before・After", "", "ステータス"]
    _row(ws, row, [(l, C_HEADER, True) for l in labels], row_h=20)


def add_ur_row(ws, row, ur_id, title, reason, explanation="", status=""):
    """UR 3行セット（縦配置）。次の行番号を返す。"""
    _row(ws, row,
         [("【ユーザ要求】", C_UR, True),
          (ur_id,           C_UR, True),
          (title,           C_UR, True),
          ("",              C_UR, True),
          ("",              C_UR, True),
          (status or "",    C_UR, True)],
         row_h=40)
    _row(ws, row + 1,
         [("",           C_UR, True),
          ("理由",       C_UR, True),
          (reason or "", C_UR, True),
          ("",           C_UR, True),
          ("",           C_UR, True),
          ("",           C_UR, True)])
    _row(ws, row + 2,
         [("",                  C_UR, True),
          ("説明",              C_UR, True),
          (explanation or "",   C_UR, True),
          ("",                  C_UR, True),
          ("",                  C_UR, True),
          ("",                  C_UR, True)])
    return row + 3


def add_sr_row(ws, row, sr_id, title, reason, explanation="", status=""):
    """SR 3行セット（縦配置）。次の行番号を返す。"""
    _row(ws, row,
         [("【システム要求】", C_SR, True),
          ("",          C_SR, True),
          (sr_id,       C_SR, True),
          (title,       C_SR, True),
          ("",          C_SR, True),
          (status or "", C_SR, True)],
         row_h=40)
    _row(ws, row + 1,
         [("",           C_SR, True),
          ("",           C_SR, True),
          ("理由",       C_SR, True),
          (reason or "", C_SR, True),
          ("",           C_SR, True),
          ("",           C_SR, True)])
    _row(ws, row + 2,
         [("",                  C_SR, True),
          ("",                  C_SR, True),
          ("説明",              C_SR, True),
          (explanation or "",   C_SR, True),
          ("",                  C_SR, True),
          ("",                  C_SR, True)])
    return row + 3


def add_sp_rows(ws, start_row, sp_id, title, before, after, biko="", kenen="", status=""):
    """SP 3〜5行セット。次の行番号を返す。"""
    r = start_row

    _row(ws, r,
         [("",      C_SP, False),
          ("",      C_SP, False),
          (title,   C_SP, False),
          ("",      C_SP, False),
          ("",      C_SP, False),
          ("",      C_SP, False)],
         wrap=False)
    r += 1

    _row(ws, r,
         [("【仕様】",       C_BEFORE, False),
          ("",               C_BEFORE, False),
          (sp_id,            C_BEFORE, False),
          ("■ Before",       C_BEFORE, False),
          (before or "",     C_BEFORE, False),
          (status or "",     C_BEFORE, False)],
         row_h=45)
    r += 1

    _row(ws, r,
         [("",               C_AFTER, False),
          ("",               C_AFTER, False),
          ("",               C_AFTER, False),
          ("■ After",        C_AFTER, False),
          (after or "",      C_AFTER, False),
          ("",               C_AFTER, False)],
         row_h=45)
    r += 1

    if biko:
        _row(ws, r,
             [("",           C_BIKO, False),
              ("",           C_BIKO, False),
              ("",           C_BIKO, False),
              ("■ 備考",     C_BIKO, False),
              (biko,         C_BIKO, False),
              ("",           C_BIKO, False)])
        r += 1

    if kenen:
        _row(ws, r,
             [("",                   C_KENEN, False),
              ("",                   C_KENEN, False),
              ("",                   C_KENEN, False),
              ("■ 懸念・検討事項",   C_KENEN, False),
              (kenen,                C_KENEN, False),
              ("",                   C_KENEN, False)])
        r += 1

    return r


def add_pending_section(ws, row, items):
    """未決事項セクション。items = list of (number, title, content, deadline)。次の行番号を返す。"""
    r = row
    _row(ws, r,
         [("■ 未決事項", C_PEND_H, True),
          ("#",           C_PEND_H, True),
          ("項目",        C_PEND_H, True),
          ("内容",        C_PEND_H, True),
          ("対応期限",    C_PEND_H, True),
          ("",            C_PEND_H, True)],
         row_h=20)
    r += 1
    for num, title, content, deadline in items:
        _row(ws, r,
             [("",           C_PEND, False),
              (str(num),     C_PEND, False),
              (title or "",  C_PEND, False),
              (content or "",C_PEND, False),
              (deadline or "",C_PEND, False),
              ("",           C_PEND, False)])
        r += 1
    return r


def add_notes_section(ws, row, items):
    """気づき・提案メモセクション。items = list of (number, kind, content, policy)。次の行番号を返す。"""
    r = row
    _row(ws, r,
         [("■ 気づき・提案メモ", C_NOTES_H, True),
          ("#",                  C_NOTES_H, True),
          ("種別",               C_NOTES_H, True),
          ("内容",               C_NOTES_H, True),
          ("対応方針",           C_NOTES_H, True),
          ("",                   C_NOTES_H, True)],
         row_h=20)
    r += 1
    for num, kind, content, policy in items:
        _row(ws, r,
             [("",           C_NOTES, False),
              (str(num),     C_NOTES, False),
              (kind or "",   C_NOTES, False),
              (content or "",C_NOTES, False),
              (policy or "", C_NOTES, False),
              ("",           C_NOTES, False)])
        r += 1
    return r


def add_scope_out_section(ws, row, items):
    """スコープ外事項セクション。items = list of (number, target, reason, cr_text)。次の行番号を返す。"""
    r = row
    _row(ws, r,
         [("■ スコープ外事項", C_SCOPE_H, True),
          ("#",                C_SCOPE_H, True),
          ("対象",             C_SCOPE_H, True),
          ("除外理由",         C_SCOPE_H, True),
          ("CR原文",           C_SCOPE_H, True),
          ("",                 C_SCOPE_H, True)],
         row_h=20)
    r += 1
    for num, target, reason, cr_text in items:
        _row(ws, r,
             [("",              C_SCOPE, False),
              (str(num),        C_SCOPE, False),
              (target or "",    C_SCOPE, False),
              (reason or "",    C_SCOPE, False),
              (cr_text or "",   C_SCOPE, False),
              ("",              C_SCOPE, False)])
        r += 1
    return r


def add_impl_ref_section(ws, row, items):
    """前提条件・実装参考情報セクション。items = list of (number, kind, content, cr_text)。次の行番号を返す。"""
    r = row
    _row(ws, r,
         [("■ 前提条件・実装参考情報", C_IMPL_REF_H, True),
          ("#",                        C_IMPL_REF_H, True),
          ("種別",                     C_IMPL_REF_H, True),
          ("内容",                     C_IMPL_REF_H, True),
          ("CR原文",                   C_IMPL_REF_H, True),
          ("",                         C_IMPL_REF_H, True)],
         row_h=20)
    r += 1
    for num, kind, content, cr_text in items:
        _row(ws, r,
             [("",              C_IMPL_REF, False),
              (str(num),        C_IMPL_REF, False),
              (kind or "",      C_IMPL_REF, False),
              (content or "",   C_IMPL_REF, False),
              (cr_text or "",   C_IMPL_REF, False),
              ("",              C_IMPL_REF, False)])
        r += 1
    return r


def set_column_widths(ws):
    ws.column_dimensions['A'].width = 14.0
    ws.column_dimensions['B'].width = 13.0
    ws.column_dimensions['C'].width = 32.0
    ws.column_dimensions['D'].width = 48.0
    ws.column_dimensions['E'].width = 37.5
    ws.column_dimensions['F'].width = 18.0


def add_history_sheet(wb, history_rows):
    """history_rows: list of (version, date, author, description)"""
    ws = wb.create_sheet("変更履歴")
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 65

    for col, h in enumerate(["版数", "日付", "変更者", "変更内容"], 1):
        c = ws.cell(row=1, column=col)
        c.value = h
        c.fill  = _fill(C_HEADER)
        c.font  = Font(bold=True, color="FFFFFF")
        c.alignment = _al()
        c.border = _bdr()
    ws.row_dimensions[1].height = 20

    for ri, (ver, date, author, desc) in enumerate(history_rows, 2):
        for col, val in enumerate([ver, date, author, desc], 1):
            c = ws.cell(row=ri, column=col)
            c.value = val
            c.fill  = _fill(C_SP)
            c.font  = Font(bold=False, color="000000")
            c.alignment = _al()
            c.border = _bdr()
        ws.row_dimensions[ri].height = 50


# ── Markdown Parser ──────────────────────────────────────────────────────────

@dataclass
class SPItem:
    sp_id: str
    title: str
    status: str = ""
    before: str = ""
    after: str = ""
    biko: str = ""
    kenen: str = ""


@dataclass
class SRItem:
    sr_id: str
    title: str
    status: str = ""
    reason: str = ""
    explanation: str = ""
    sp_list: List[SPItem] = field(default_factory=list)


@dataclass
class URItem:
    ur_id: str
    title: str
    status: str = ""
    reason: str = ""
    explanation: str = ""
    sr_list: List[SRItem] = field(default_factory=list)


def _get_field(line, *markers):
    """Extract the value after '- **marker：** ' or '- **marker:** '. Returns None if not matched."""
    for marker in markers:
        for sep in ('：', ':'):
            m = re.match(rf'^\s*-\s+\*\*{re.escape(marker)}{sep}\*\*\s*(.*)', line)
            if m:
                return m.group(1).strip()
    return None


def _parse_table(lines):
    """Parse a Markdown table from a list of lines.
    Returns data rows (list of cell-string lists), skipping the header and separator rows.
    """
    rows = []
    phase = 'header'
    for line in lines:
        s = line.strip()
        if not s.startswith('|'):
            continue
        if phase == 'header':
            phase = 'sep'
            continue
        if phase == 'sep':
            phase = 'data'
            continue
        cols = [c.strip() for c in s.split('|')[1:-1]]
        if cols:
            rows.append(cols)
    return rows


def parse_crs_md(md_path: str) -> dict:
    """CRS Markdown を読み込み、UR/SR/SP 階層と付記セクションを構造化して返す。

    Return dict keys:
      urs      : list[URItem]  — UR/SR/SP の3層ネスト構造
      pending  : list[tuple]   — (no, title, content, deadline)
      notes    : list[tuple]   — (no, kind, content, policy)
      scope_out: list[tuple]   — (no, target, reason, cr_text)
      impl_ref : list[tuple]   — (no, kind, content, cr_text)
      history  : list[tuple]   — (版数, 日付, 変更者, 変更内容)
    """
    with open(md_path, encoding='utf-8') as f:
        lines = f.readlines()

    urs: List[URItem] = []
    pending = []
    notes = []
    scope_out = []
    impl_ref = []
    history = []

    cur_ur: URItem = None
    cur_sr: SRItem = None
    cur_sp: SPItem = None
    section = None  # 'usdm' | 'pending' | 'notes' | 'scope_out' | 'impl_ref' | 'history'

    i = 0
    while i < len(lines):
        stripped = lines[i].rstrip('\n')

        # ── Section heading detection (## level only) ──────────────────────
        if re.match(r'^## ', stripped):
            if re.match(r'^## 付記A\.', stripped):
                section = 'scope_out'
            elif re.match(r'^## 付記B\.', stripped):
                section = 'impl_ref'
            elif '要求仕様' in stripped or re.match(r'^## 2\.', stripped):
                section = 'usdm'
            elif '未決' in stripped:
                section = 'pending'
            elif '気づき' in stripped or '提案メモ' in stripped:
                section = 'notes'
            elif '変更履歴' in stripped:
                section = 'history'
            else:
                section = None
            i += 1
            continue

        # ── USDM hierarchy parsing ─────────────────────────────────────────
        if section == 'usdm':
            # UR heading (h5): ##### UR-xxx or ##### UR-NF-xxx
            m = re.match(r'^##### (UR-\S+)\s+(.*)', stripped)
            if m:
                cur_ur = URItem(ur_id=m.group(1), title=m.group(2).strip())
                urs.append(cur_ur)
                cur_sr = None
                cur_sp = None
                i += 1
                continue

            # SR heading (h6): ###### SR-xxx-yyy or ###### SR-NF-xxx-yyy
            m = re.match(r'^###### (SR-\S+)\s+(.*)', stripped)
            if m:
                cur_sr = SRItem(sr_id=m.group(1), title=m.group(2).strip())
                if cur_ur is not None:
                    cur_ur.sr_list.append(cur_sr)
                cur_sp = None
                i += 1
                continue

            # SP heading (h7): ####### SP-xxx-yyy.zzz or ####### SP-NF-xxx-yyy.zzz
            m = re.match(r'^####### (SP-\S+)\s+(.*)', stripped)
            if m:
                cur_sp = SPItem(sp_id=m.group(1), title=m.group(2).strip())
                if cur_sr is not None:
                    cur_sr.sp_list.append(cur_sp)
                i += 1
                continue

            # Field lines
            if cur_sp is not None:
                for attr, markers in [
                    ('before',  ('Before',)),
                    ('after',   ('After',)),
                    ('biko',    ('備考',)),
                    ('kenen',   ('懸念・検討事項',)),
                    ('status',  ('ステータス',)),
                ]:
                    v = _get_field(stripped, *markers)
                    if v is not None:
                        setattr(cur_sp, attr, v)
                        break
            elif cur_sr is not None:
                for attr, markers in [
                    ('reason',      ('理由',)),
                    ('explanation', ('説明',)),
                    ('status',      ('ステータス',)),
                ]:
                    v = _get_field(stripped, *markers)
                    if v is not None:
                        setattr(cur_sr, attr, v)
                        break
            elif cur_ur is not None:
                for attr, markers in [
                    ('reason',      ('理由',)),
                    ('explanation', ('説明',)),
                    ('status',      ('ステータス',)),
                ]:
                    v = _get_field(stripped, *markers)
                    if v is not None:
                        setattr(cur_ur, attr, v)
                        break

        # ── Table section parsing ──────────────────────────────────────────
        elif section in ('pending', 'notes', 'scope_out', 'impl_ref', 'history'):
            if stripped.strip().startswith('|'):
                # Collect all consecutive table lines
                j = i
                table_lines = []
                while j < len(lines) and lines[j].strip().startswith('|'):
                    table_lines.append(lines[j])
                    j += 1
                rows = _parse_table(table_lines)
                if section == 'pending':
                    for row in rows:
                        if len(row) >= 4:
                            pending.append((row[0], row[1], row[2], row[3]))
                elif section == 'notes':
                    for row in rows:
                        if len(row) >= 4:
                            notes.append((row[0], row[1], row[2], row[3]))
                elif section == 'scope_out':
                    for row in rows:
                        if len(row) >= 4:
                            scope_out.append((row[0], row[1], row[2], row[3]))
                elif section == 'impl_ref':
                    for row in rows:
                        if len(row) >= 4:
                            impl_ref.append((row[0], row[1], row[2], row[3]))
                elif section == 'history':
                    for row in rows:
                        if len(row) >= 4:
                            history.append((row[0], row[1], row[2], row[3]))
                i = j
                continue

        i += 1

    return {
        'urs':       urs,
        'pending':   pending,
        'notes':     notes,
        'scope_out': scope_out,
        'impl_ref':  impl_ref,
        'history':   history,
    }


# ── Top-level builder ────────────────────────────────────────────────────────

def build_excel_from_md(md_path: str, out_path: str) -> None:
    """CRS Markdown を Excel に変換して out_path に保存する。"""
    data = parse_crs_md(md_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "変更要求仕様書"
    set_column_widths(ws)

    r = 1
    add_header_row(ws, r)
    r += 1

    for ur in data['urs']:
        r = add_ur_row(ws, r, ur.ur_id, ur.title, ur.reason, ur.explanation, ur.status)
        for sr in ur.sr_list:
            r = add_sr_row(ws, r, sr.sr_id, sr.title, sr.reason, sr.explanation, sr.status)
            for sp in sr.sp_list:
                r = add_sp_rows(ws, r, sp.sp_id, sp.title, sp.before, sp.after,
                                sp.biko, sp.kenen, sp.status)

    if data['pending']:
        r = add_pending_section(ws, r, data['pending'])
    if data['notes']:
        r = add_notes_section(ws, r, data['notes'])
    # スコープ外・実装参考情報はリストが空でもヘッダ行のみ出力する
    r = add_scope_out_section(ws, r, data['scope_out'])
    r = add_impl_ref_section(ws, r, data['impl_ref'])

    if data['history']:
        add_history_sheet(wb, data['history'])

    wb.save(out_path)
    print(f"Generated: {out_path}  (rows: {r - 1})")


# ── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit("Usage: python crs_md2excel.py <CRS_MD_PATH> <OUTPUT_XLSX_PATH>")
    md_path, out_path = sys.argv[1], sys.argv[2]
    if not out_path.lower().endswith(".xlsx"):
        sys.exit(f"Error: OUTPUT_XLSX_PATH must end with .xlsx — got: {out_path!r}")
    if not os.path.exists(md_path):
        sys.exit(f"Error: Markdown file not found: {md_path!r}")
    build_excel_from_md(md_path, out_path)
