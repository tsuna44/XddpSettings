import sys
import tempfile
import unittest
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

if openpyxl is not None:
    from crs_md2excel import build_excel_from_md, parse_crs_md  # noqa: E402

UR_BLOCK_START = "【ユーザ要求】"  # add_ur_row()が全UR行セット（実UR・プレースホルダ共通）の
                                    # 1行目・列Aに出力する固定ラベル。行ブロックの境界マーカーとして使う。

CRS_TEXT_NORMAL = """# 変更要求仕様書

## 2. USDM 要求仕様

### ＜機能要求＞

#### CR-2026-970-UR-001 通常のUR

##### ＜要求グループ＞

###### CR-2026-970-SR-001-001 通常のSR

**＜仕様グループ＞**

- **CR-2026-970-SP-001-001.001**: 通常のSP

## 3. トレーサビリティマトリクス（TM）
"""

CRS_TEXT_ORPHAN_SR = """# 変更要求仕様書

## 2. USDM 要求仕様

### ＜機能要求＞

#### CR-2026-970-UR-001 通常のUR

##### ＜要求グループ＞

###### CR-2026-970-SR-001-001 通常のSR

**＜仕様グループ＞**

- **CR-2026-970-SP-001-001.001**: 通常のSP

#### 本CRの対象範囲を限定する（形式的なURを持たないスコープ宣言）

###### CR-2026-970-SR-999-001 スコープ除外の宣言（親URなし）

## 3. トレーサビリティマトリクス（TM）
"""


CRS_TEXT_WITH_REASON = """# 変更要求仕様書

## 2. USDM 要求仕様

### ＜機能要求＞

#### CR-2026-970-UR-001 通常のUR

##### ＜要求グループ＞

###### CR-2026-970-SR-001-001 通常のSR

**＜仕様グループ＞**

- **CR-2026-970-SP-001-001.001**: 通常のSP
  - **Before：** 旧処理をする
  - **After：** 新処理をする
  - **理由：** 保守性向上のため
  - **備考：** 制約なし

## 3. トレーサビリティマトリクス（TM）
"""


CRS_TEXT_WITH_SPEC_ONLY = """# 変更要求仕様書

## 2. USDM 要求仕様

### ＜機能要求＞

#### CR-2026-970-UR-001 新規開発のUR

##### ＜要求グループ＞

###### CR-2026-970-SR-001-001 新規開発のSR

**＜仕様グループ＞**

- **CR-2026-970-SP-001-001.001**: 新規SP
  - **仕様：** 新しい処理を実装する
  - **備考：** 制約なし

## 3. トレーサビリティマトリクス（TM）
"""


@unittest.skipIf(openpyxl is None, "openpyxl not installed")
class CrsMd2ExcelTestCase(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tmpdir.name)

    def tearDown(self):
        self.tmpdir.cleanup()

    def _build(self, text):
        md_path = self.root / "CRS-TEST.md"
        md_path.write_text(text, encoding="utf-8")
        out_path = self.root / "out.xlsx"
        build_excel_from_md(str(md_path), str(out_path))
        wb = openpyxl.load_workbook(out_path)
        return wb.active

    def test_normal_ur_sr_sp_hierarchy_is_parsed(self):
        md_path = self.root / "CRS-TEST.md"
        md_path.write_text(CRS_TEXT_NORMAL, encoding="utf-8")
        data = parse_crs_md(str(md_path))
        self.assertEqual(len(data["urs"]), 1)
        ur = data["urs"][0]
        self.assertEqual(ur.ur_id, "CR-2026-970-UR-001")
        self.assertEqual(len(ur.sr_list), 1)
        sr = ur.sr_list[0]
        self.assertEqual(sr.sr_id, "CR-2026-970-SR-001-001")
        self.assertEqual(len(sr.sp_list), 1)
        self.assertEqual(sr.sp_list[0].sp_id, "CR-2026-970-SP-001-001.001")

    def test_requirement_group_h5_does_not_create_ghost_ur(self):
        # 新体系では要求グループが H5（##### ＜…＞）。旧 h5 フォールバックが残っていると
        # 要求グループ見出しを幽霊 UR 行として Excel に出力する回帰（#10）を検出する。
        md_path = self.root / "CRS-TEST.md"
        md_path.write_text(CRS_TEXT_NORMAL, encoding="utf-8")  # 要求グループ ＜要求グループ＞ を含む
        data = parse_crs_md(str(md_path))
        self.assertEqual(len(data["urs"]), 1)
        self.assertEqual(data["urs"][0].ur_id, "CR-2026-970-UR-001")
        # 要求グループ名が UR タイトルとして混入していないこと
        self.assertNotIn("要求グループ", data["urs"][0].title)

    def test_sr_without_parent_ur_is_not_misattached(self):
        ws = self._build(CRS_TEXT_ORPHAN_SR)

        ur001_row = next(
            r for r in range(1, ws.max_row + 1)
            if ws.cell(r, 1).value == UR_BLOCK_START and ws.cell(r, 2).value == "CR-2026-970-UR-001"
        )
        # CR-2026-970-UR-001の次のUR行セット（実UR・プレースホルダいずれか。またはシート末尾）までの
        # 範囲にCR-2026-970-SR-999-001が出現しないこと
        next_ur_block_row = next(
            (r for r in range(ur001_row + 1, ws.max_row + 1)
             if ws.cell(r, 1).value == UR_BLOCK_START),
            ws.max_row + 1,
        )
        sr_ids_under_ur001 = [
            ws.cell(r, 3).value for r in range(ur001_row + 1, next_ur_block_row)
        ]
        self.assertNotIn("CR-2026-970-SR-999-001", sr_ids_under_ur001)

        # CR-2026-970-SR-999-001自体はどこかに出力されている（サイレントに消えていない）こと
        all_sr_ids = [ws.cell(r, 3).value for r in range(1, ws.max_row + 1)]
        self.assertIn("CR-2026-970-SR-999-001", all_sr_ids)

        # 親URなしグループの見出し行が、CR-2026-970-UR-001とは別の独立したUR行セットとして
        # 出力されていること（＝サイレントにCR-2026-970-UR-001へ吸収されていないことの直接確認）
        placeholder_row = next(
            r for r in range(1, ws.max_row + 1)
            if ws.cell(r, 1).value == UR_BLOCK_START
            and ws.cell(r, 2).value in (None, "")
            and ws.cell(r, 3).value and "スコープ宣言" in ws.cell(r, 3).value
        )
        self.assertGreater(placeholder_row, ur001_row)


    def test_sp_reason_is_parsed(self):
        md_path = self.root / "CRS-TEST.md"
        md_path.write_text(CRS_TEXT_WITH_REASON, encoding="utf-8")
        data = parse_crs_md(str(md_path))
        sp = data["urs"][0].sr_list[0].sp_list[0]
        self.assertEqual(sp.reason, "保守性向上のため")
        self.assertEqual(sp.biko, "制約なし")

    def test_sp_reason_row_between_after_and_biko(self):
        ws = self._build(CRS_TEXT_WITH_REASON)
        # D列（■ ラベル）の出現順を収集
        d_labels = [ws.cell(r, 4).value for r in range(1, ws.max_row + 1)]
        after_idx = d_labels.index("■ After")
        reason_idx = d_labels.index("■ 理由")
        biko_idx = d_labels.index("■ 備考")
        self.assertLess(after_idx, reason_idx)
        self.assertLess(reason_idx, biko_idx)
        # 理由行のE列に本文が出力されていること
        reason_row = reason_idx + 1  # d_labels は0始まり・行番号は1始まり
        self.assertEqual(ws.cell(reason_row, 5).value, "保守性向上のため")

    def test_sp_spec_only_is_parsed(self):
        md_path = self.root / "CRS-TEST.md"
        md_path.write_text(CRS_TEXT_WITH_SPEC_ONLY, encoding="utf-8")
        data = parse_crs_md(str(md_path))
        sp = data["urs"][0].sr_list[0].sp_list[0]
        self.assertEqual(sp.spec, "新しい処理を実装する")
        self.assertEqual(sp.before, "")
        self.assertEqual(sp.after, "")

    def test_sp_spec_only_outputs_single_spec_row(self):
        ws = self._build(CRS_TEXT_WITH_SPEC_ONLY)
        d_labels = [ws.cell(r, 4).value for r in range(1, ws.max_row + 1)]
        # Before/After 行は出力されず、仕様行のみが出力される
        self.assertNotIn("■ Before", d_labels)
        self.assertNotIn("■ After", d_labels)
        spec_idx = d_labels.index("■ 仕様")
        spec_row = spec_idx + 1
        self.assertEqual(ws.cell(spec_row, 5).value, "新しい処理を実装する")
        # 備考行も引き続き出力される
        biko_idx = d_labels.index("■ 備考")
        self.assertLess(spec_idx, biko_idx)


if __name__ == "__main__":
    unittest.main()
